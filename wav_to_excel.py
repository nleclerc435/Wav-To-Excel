import wave
import contextlib
import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Border, Side
import os
import glob

#Generic cell formating for the header row
def format_cells(range):
    for cell in ws[range]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('b2b2b2'))
        cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             top=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'))

#Creating a new workbook and setting up the header row.
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'File Name'
ws['B1'] = 'Duration'
ws['C1'] = 'Description'
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 100
format_cells('1:1')

while True:
    path = raw_input('Please, enter the path of the folder where your .wav files are located.\n').strip('\"')
    if os.path.exists(path):
        #Adding the backslash if it is missing to have a valid path to the file
        if path[-1] != "\\" :
            path = path + "\\"
        items = [os.path.basename(file) for file in glob.glob(os.path.join(path, '*.wav'))]
        #Making sure there are wav files in the directory. If not, user is prompted to enter path again.
        if len(items) == 0:
            print '\nNo wav files found. Are you sure you entered the right path?\n'
        else:
            count = 2
            for i in items:
                #Open each .wav files and get the number of frames, framerate to determine the duration. Using divmod to get the time in minutes:seconds instead of seconds only.
                with contextlib.closing(wave.open(path+i, 'r')) as f:
                    print 'Adding {} ...'.format(i)
                    frames = f.getnframes()
                    rate = f.getframerate()
                    duration = frames/float(rate)
                    m,s = divmod(duration, 60)
                    ws['A{}'.format(count)] = i
                    ws['B{}'.format(count)] = '{:02.0f}m{:02.0f}s'.format(m, s)
                    count += 1
            print 'Done!'
            wb.save('wav_to_excel.xlsx')
            break
    # If path does not exists, loop back to path request.
    else:
        print 'Invalid path. Please make sure you entered the correct path.'
        continue





