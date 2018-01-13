import Tkinter, tkFileDialog, tkMessageBox
import glob
import os
import contextlib
import wave
import openpyxl
from openpyxl.styles import Font, colors
import sys
import lbwmenu
import os

class App():

    def __init__(self, master):

        frame = Tkinter.Frame(master)
        frame.pack()
        
        #Send to Excel button
        self.exl_but = Tkinter.Button(frame, text='Send to Excel', command=self.send_to_excel)
        self.exl_but.pack(side='right', padx=10, pady=5)

        #Path text
        self.label = Tkinter.Label(frame, text='Path: ')
        self.label.pack(fill='x')

        #Listbox with scrollbars
        self.v_scrollbar = Tkinter.Scrollbar(frame, orient='vertical')
        self.h_scrollbar = Tkinter.Scrollbar(frame, orient='horizontal')
        self.lb = lbwmenu.ListboxWMenu(frame, width=100, height=20, yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set, selectmode='extended')
        self.v_scrollbar.config(command=self.lb.yview)
        self.h_scrollbar.config(command=self.lb.xview)
        self.h_scrollbar.pack(side='bottom', fill='x')
        self.v_scrollbar.pack(side='right', fill='y')
        self.lb.pack(padx=5, pady=5, expand=True)

        #Menu bar
        self.menubar = Tkinter.Menu(root)
        root.config(menu=self.menubar)
        self.filemenu = Tkinter.Menu(self.menubar, tearoff=False)
        self.menubar.add_cascade(label='File', menu=self.filemenu)
        self.filemenu.add_command(label='Open', command=self.open_dialog, accelerator='Ctrl+O')
        self.filemenu.add_separator()
        self.filemenu.add_command(label='Quit', command=self.quit, accelerator='Ctrl+Q')

        #Keyboard shorcuts
        root.bind_all('<Control-o>', self.open_dialog)
        root.bind_all('<Control-q>', self.quit)

    def delete_entry(self):
        for entry in self.lb.curselection()[::-1]:
            self.lb.delete(entry)

    def quit(self, *args):
        sys.exit(0)

    def get_duration(self, list, path):
        duration_list = []
        for i in list:
            with contextlib.closing(wave.open(path+i, 'r')) as f:
                frames = f.getnframes()
                rate = f.getframerate()
                duration = frames/float(rate)
                m,s = divmod(duration, 60)
                duration_list.append('{:02.0f}m{:02.0f}s'.format(m, s))
        return duration_list    


    def open_dialog(self, *args):
        path = str(tkFileDialog.askdirectory())
        if path != '':
            cleaned_path = path.replace('/', '\\') + '\\'
            self.label['text'] = 'Path: ' + cleaned_path
            self.lb.delete(0, 'end')
            items = [os.path.basename(file) for file in glob.glob(os.path.join(cleaned_path, '*.wav'))]
            durations = self.get_duration(items, cleaned_path)
            count = 1
            for i in range(len(items)):
                self.lb.insert(count, items[i] + '    Duration: ' + durations[i])
                count += 1


    def setup_wb(self,wb):
        ws = wb.active
        ws['A1'] = 'File Name'
        ws['B1'] = 'Duration'
        ws['C1'] = 'Description'
        ws['D1'] = 'Hyperlink'
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 100
        ws.column_dimensions['D'].width = 20


    def send_to_excel(self):
        if self.lb.size() > 0:
            self.lb.select_set(0,'end')
            items = [self.lb.get(i) for i in self.lb.curselection()]
            try:
                filename = tkFileDialog.asksaveasfilename(defaultextension='.xlsx',filetypes=(('Excel files','*.xlsx'), ('all files','*.*')))
                if filename != None or filename != '':
                    wb = openpyxl.Workbook()
                    self.setup_wb(wb)
                    ws = wb.active
                    names = [name.split('    Duration: ')[0] for name in items]
                    durations = [duration.split('    Duration: ')[1] for duration in items]
                    for i in range(0,len(items)):
                        ws['A{}'.format(i+2)] = names[i]
                        ws['B{}'.format(i+2)] = durations[i]
                        ws['D{}'.format(i+2)].hyperlink = self.label['text'].split('Path: ')[1] + names[i]
                        ws['D{}'.format(i+2)] = 'LINK'
                        ws['D{}'.format(i+2)].font = Font(bold=True,underline='single', color=colors.BLUE)
                    wb.save(filename)
                    self.lb.select_clear(0, 'end')
                    if tkMessageBox.askyesno('Success', 'Done! \nDo you want to open the Excel file?'):
                        os.startfile(filename)
            except IOError:
                pass    
        else:
            tkMessageBox.showerror('Error', 'No data to transfer.')

            

root = Tkinter.Tk()
root.geometry('730x400')
root.wm_title('Wav2Excel')
root.iconbitmap('python_ico.ico')
root.resizable(0,0)
app = App(root)
root.mainloop()